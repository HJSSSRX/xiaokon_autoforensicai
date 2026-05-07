"""
sync_kb.py — 知识库归档同步工具

把比赛/训练的实时协作数据（Hub 里的 findings/answers/progress）
同步到 knowledge/competitions/<赛事名>/ 下，让人类能浏览实时进度。

【设计原则】
- knowledge/ = 永久资产（人类阅读 + 长期检索）
- case/shared/ = 临时工作（Hub 机读）
- sync_kb.py = 桥梁，把临时数据归档为永久 markdown

【输出文件】
- MASTER_SHEET.md  实时主表（按类别分组）
- findings_snapshot.yaml  完整 finding 快照
- 不动 README.md 和 evidence_index.md（手工维护）

Usage:
  python tools/sync_kb.py [--hub URL] [--out KB_DIR]

  默认从 http://127.0.0.1:8765 拉数据，写到
  knowledge/competitions/2026FIC-团体赛/

注意：这个脚本会**覆盖**输出文件中由它生成的部分。
手工编辑过的 README/evidence_index 不会被动到。
MASTER_SHEET.md 整个文件会被重写。
"""
import argparse
import datetime
import json
import sys
import urllib.error
import urllib.request
from pathlib import Path

import yaml

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    HAVE_XLSX = True
except ImportError:
    HAVE_XLSX = False


# ─────────────────────────────────────────────────────
# 题目元数据 — 写死在脚本里（题目原文不变）
# ─────────────────────────────────────────────────────

QUESTIONS = {
    "mobile_forensics": {
        "title": "手机取证",
        "items": [
            ("Q1",  "该手机型号"),
            ("Q2",  "李安弘计划前往迪拜的日期"),
            ("Q3",  "与网站搭建人员沟通的 APP 安装日期"),
            ("Q4",  "该 APP 聊天数据库文件名"),
            ("Q5",  "数据库解密密码（SQLCipher）"),
            ("Q6",  "云服务器商家备用钱包地址"),
            ("Q7",  "给搭建人员第一次转账 hash 前 6 位"),
            ("Q8",  "主动向 AI 提问次数"),
            ("Q9",  "AI 软件调用本地 AI 模型版本"),
            ("Q10", "无人机飞行县/区"),
            ("Q11", "视频类 APP 涉敏感权限（多选）"),
            ("Q12", "网络断开时加载的本地离线页面"),
            ("Q13", "数据上传地址（解码后）"),
            ("Q14", "存用户信息的 SQLite 表名"),
            ("Q15", "assets/config.dat 解密后的 USDT 钱包"),
            ("Q16", "JS Bridge 暴露的获取通讯录方法"),
            ("Q17", "备用服务器完整域名:端口"),
        ],
    },
    "binary_forensics": {
        "title": "二进制取证",
        "items": [
            ("Q1", "SampleVC.exe MD5"),
            ("Q2", "编译日期"),
            ("Q3", "正确的密码"),
            ("Q4", "解密后文件后缀"),
            ("Q5", "李安弘虚拟币收款总金额"),
        ],
    },
    "server_forensics": {
        "title": "服务器取证",
        "items": [
            ("Q1",  "操作系统版本"),
            ("Q2",  "根分区硬盘 UUID"),
            ("Q3",  "最新 docker 镜像创建时间"),
            ("Q4",  "根分区快照路径"),
            ("Q5",  "网站后台管理入口文件名"),
            ("Q6",  "网站 ICP 备案号"),
            ("Q7",  "网站主域名"),
            ("Q8",  "分类3视频拼音"),
            ("Q9",  "站点设置页面前端模板源文件"),
            ("Q10", "伪静态规则文件 SM3"),
            ("Q11", "关联数据库 IP"),
            ("Q12", "数据库容器技术"),
            ("Q13", "4000 端口备份数据库版本"),
            ("Q14", "新注册用户数最多的日期"),
            ("Q15", "马慧美最后登录 IP"),
            ("Q16", "未被使用的文件系统（A/B/C/D）"),
            ("Q17", "已安装数据库（多选）"),
        ],
    },
    "internet_forensics": {
        "title": "互联网取证",
        "items": [
            ("Q1", "售卖卡密的公开群组 ID"),
            ("Q2", "备份数据库视频图片文件名"),
            ("Q3", "ngrok 提供的域名"),
        ],
    },
    "computer_forensics": {
        "title": "计算机取证",
        "items": [
            ("Q1",  "操作系统版本号"),
            ("Q2",  "钓鱼邮件发送方邮箱"),
            ("Q3",  "黄金换现金商家联系方式"),
            ("Q4",  "推广设计图中 apk 下载链接"),
            ("Q5",  "VPN 软件代理端口"),
            ("Q6",  "AI 软件当前模型类型"),
            ("Q7",  "AI 软件 apiKey"),
            ("Q8",  "勒索软件解密服务联系方式"),
            ("Q9",  "存放黄金的保险柜编号"),
            ("Q10", "保险柜密码"),
        ],
    },
}

CATEGORY_OWNER = {
    "mobile_forensics":   "mobile_analyst",
    "binary_forensics":   "binary_analyst",
    "server_forensics":   "server_analyst",
    "internet_forensics": "server_analyst",
    "computer_forensics": "computer_analyst",
}


# ─────────────────────────────────────────────────────
# Hub HTTP client
# ─────────────────────────────────────────────────────

def get_json(url):
    try:
        with urllib.request.urlopen(url, timeout=10) as r:
            return json.loads(r.read().decode("utf-8"))
    except urllib.error.URLError as e:
        sys.exit(f"Hub 不可达 {url}: {e}")


def fetch_hub_state(hub):
    return {
        "ping":     get_json(f"{hub}/ping"),
        "findings": get_json(f"{hub}/findings"),
        "answers":  get_json(f"{hub}/answers"),
        "progress": get_json(f"{hub}/progress"),
        "session":  get_json(f"{hub}/session"),
    }


# ─────────────────────────────────────────────────────
# 渲染 MASTER_SHEET.md
# ─────────────────────────────────────────────────────

def status_emoji(answer_entry):
    if not (answer_entry and answer_entry.get("answer")):
        return "[ ]"  # 未答
    vs = (answer_entry.get("verification_status") or "unverified").lower()
    if vs == "verified":
        return "[V]"  # 已验证
    if vs == "disputed":
        return "[?]"  # 争议
    if vs == "failed":
        return "[X]"  # 验证失败
    return "[Y]"      # 已答未验证


def render_master_sheet(state):
    """从 Hub state 渲染整个 MASTER_SHEET 内容（一次性覆盖写入）。"""
    answers = state["answers"] or {}
    findings = state["findings"] or []
    progress = state["progress"] or {}
    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # 索引：finding id -> finding
    findings_by_id = {f["id"]: f for f in findings}
    # 索引：answer (category, qid) -> entry
    ans_idx = {}
    for cat, lst in answers.items():
        if isinstance(lst, list):
            for a in lst:
                ans_idx[(cat, a.get("qid"))] = a

    out = []
    out.append("# MASTER_SHEET — 2026 FIC 团体赛实时主表")
    out.append("")
    out.append(f"> **最后更新**: {now} (sync_kb.py 自动生成)")
    out.append("> **更新方式**: `python tools/sync_kb.py`")
    out.append("> **图例**: [V] 已验证 · [Y] 已答未验证 · [?] 争议 · [X] 验证失败 · [ ] 未答")
    out.append("")

    # 总览
    out.append("## 总览")
    out.append("")
    out.append("| 类别 | 完成 / 总数 | 主负责角色 | 进度 |")
    out.append("|---|---|---|---|")
    total_done = total_all = 0
    for cat, meta in QUESTIONS.items():
        total = len(meta["items"])
        done = sum(1 for q, _ in meta["items"] if ans_idx.get((cat, q), {}).get("answer"))
        owner = CATEGORY_OWNER.get(cat, "?")
        st = progress.get(owner, {}).get("status", "?")
        out.append(f"| {meta['title']} | **{done} / {total}** | {owner} | {st} |")
        total_done += done
        total_all += total
    pct = (total_done * 100.0 / total_all) if total_all else 0
    out.append(f"| **合计** | **{total_done} / {total_all} ({pct:.1f}%)** | — | — |")
    out.append("")

    # 每类详情
    for cat, meta in QUESTIONS.items():
        owner = CATEGORY_OWNER.get(cat, "?")
        out.append(f"## {meta['title']} ({len(meta['items'])} 题，主答 {owner})")
        out.append("")
        out.append("| # | 题目 | 答案 | 验证 | 解析摘要 | 证据路径 | finding |")
        out.append("|---|---|---|---|---|---|---|")
        for qid, qtext in meta["items"]:
            a = ans_idx.get((cat, qid), {})
            ans = a.get("answer", "—") or "—"
            stat = status_emoji(a)
            ev_id = a.get("evidence", "") or ""
            ev_paths = a.get("evidence_path", []) or []
            analysis = a.get("analysis", "") or ""
            # 解析摘要取第一行前 80 字
            analysis_short = analysis.split("\n")[0][:80] if analysis else "—"
            # 证据路径取前 2 个
            paths_short = ", ".join(ev_paths[:2]) if ev_paths else "—"
            if len(ev_paths) > 2:
                paths_short += f" (+{len(ev_paths) - 2})"
            ev_md = f"`{ev_id}`" if ev_id else "—"

            # escape pipes in markdown table
            def md_safe(s):
                return str(s).replace("|", "\\|").replace("\n", " ")
            out.append("| {} | {} | `{}` | {} | {} | {} | {} |".format(
                qid, md_safe(qtext), md_safe(ans), stat,
                md_safe(analysis_short), md_safe(paths_short), ev_md,
            ))
        out.append("")

        # 附：该类别各题的完整解析 — 供人类复现
        has_any_analysis = any(
            ans_idx.get((cat, qid), {}).get("analysis") for qid, _ in meta["items"]
        )
        if has_any_analysis:
            out.append(f"### {meta['title']} — 详细解析")
            out.append("")
            for qid, qtext in meta["items"]:
                a = ans_idx.get((cat, qid), {})
                if not a.get("analysis"):
                    continue
                vs = a.get("verification_status", "unverified")
                vs_label = {"verified": "✓ 已验证", "unverified": "⚠ 未验证",
                            "disputed": "× 争议", "failed": "× 失败"}.get(vs, vs)
                out.append(f"#### {qid} · {qtext}")
                out.append("")
                out.append(f"**答案**: `{a.get('answer', '')}` · **验证**: {vs_label}")
                if a.get("verified_by"):
                    out.append(f"  · 验证者: {a['verified_by']} ({a.get('verified_at', '')})")
                if a.get("verify_note"):
                    out.append(f"  · 验证备注: {a['verify_note']}")
                out.append("")
                out.append("**解析**:")
                out.append("")
                # 代码块包裹以防 markdown 破坏
                analysis_text = a.get("analysis", "")
                # 如果 analysis 本身包含 ``` 则不再包
                if "```" in analysis_text:
                    out.append(analysis_text)
                else:
                    out.append(analysis_text)
                out.append("")
                ev_paths = a.get("evidence_path", []) or []
                if ev_paths:
                    out.append("**证据路径**:")
                    for p in ev_paths:
                        out.append(f"- `{p}`")
                    out.append("")
                if a.get("evidence"):
                    out.append(f"**证据 finding**: `{a['evidence']}`")
                    out.append("")

    # 跨角色 finding（无 answer 关联但 related_to 不空）
    out.append("## 跨角色关键发现")
    out.append("")
    out.append("（findings 中 `related_to` 不空、且未直接关联到某个具体 Q 的）")
    out.append("")
    cross = [f for f in findings if f.get("related_to")]
    if cross:
        out.append("| Finding ID | 来源 | 摘要 | 给谁用 |")
        out.append("|---|---|---|---|")
        for f in cross:
            related = ",".join(f.get("related_to", []))
            summary = f.get("summary", "")[:80].replace("|", "\\|")
            out.append(f"| `{f['id']}` | {f.get('from','?')} | {summary} | {related} |")
    else:
        out.append("（暂无）")
    out.append("")

    # 当前角色 progress
    out.append("## 当前各角色状态")
    out.append("")
    out.append("| 角色 | 状态 | 当前任务 | 最后更新 |")
    out.append("|---|---|---|---|")
    for role, p in progress.items():
        ct = (p.get("current_task", "") or "")[:60].replace("|", "\\|")
        out.append(f"| {role} | {p.get('status','?')} | {ct} | {p.get('updated','')} |")
    out.append("")

    # 当前阻塞
    blockers = (state["session"] or {}).get("blockers", []) or []
    open_blockers = [b for b in blockers if b.get("status") == "open"]
    if open_blockers:
        out.append("## 当前活跃阻塞")
        out.append("")
        out.append("| 阻塞 ID | 来源 | 内容 | 需要 |")
        out.append("|---|---|---|---|")
        for b in open_blockers:
            blk = b.get("blocker", "")[:60].replace("|", "\\|")
            needs = b.get("needs", "")[:60].replace("|", "\\|")
            out.append(f"| {b.get('id','?')} | {b.get('from','?')} | {blk} | {needs} |")
        out.append("")

    return "\n".join(out)


# ─────────────────────────────────────────────────────
# 渲染 MASTER_SHEET.xlsx (Excel 报告)
# ─────────────────────────────────────────────────────

def render_xlsx(state, out_path):
    """生成 Excel 答案表，列：题号|题目|答案|状态|证据 ID|证据摘要|来源角色|更新时间。"""
    if not HAVE_XLSX:
        print("[!] openpyxl not installed, skip xlsx generation")
        return
    answers = state["answers"] or {}
    findings = state["findings"] or []
    findings_by_id = {f["id"]: f for f in findings}
    ans_idx = {}
    for cat, lst in answers.items():
        if isinstance(lst, list):
            for a in lst:
                ans_idx[(cat, a.get("qid"))] = a

    wb = Workbook()
    # 删除默认空 sheet
    wb.remove(wb.active)

    # ── 总览 sheet ──
    ws_overview = wb.create_sheet("总览")
    ws_overview.append(["类别", "完成", "总数", "百分比", "主负责角色"])
    total_done = total_all = 0
    for cat, meta in QUESTIONS.items():
        total = len(meta["items"])
        done = sum(1 for q, _ in meta["items"] if ans_idx.get((cat, q), {}).get("answer"))
        owner = CATEGORY_OWNER.get(cat, "?")
        pct = f"{done * 100.0 / total:.1f}%" if total else "0%"
        ws_overview.append([meta["title"], done, total, pct, owner])
        total_done += done
        total_all += total
    pct_total = f"{total_done * 100.0 / total_all:.1f}%" if total_all else "0%"
    ws_overview.append(["合计", total_done, total_all, pct_total, "—"])

    # 表头加粗 + 上色
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws_overview[1]:
        cell.fill = header_fill
        cell.font = header_font

    # 调整列宽
    for col_letter, width in [("A", 18), ("B", 8), ("C", 8), ("D", 10), ("E", 22)]:
        ws_overview.column_dimensions[col_letter].width = width

    # ── 各类别 sheet（每个类别一个 sheet）──
    status_fill = {
        "verified":   PatternFill("solid", fgColor="C6EFCE"),
        "unverified": PatternFill("solid", fgColor="FFEB9C"),
        "disputed":   PatternFill("solid", fgColor="FFC7CE"),
        "failed":     PatternFill("solid", fgColor="D9D9D9"),
    }
    for cat, meta in QUESTIONS.items():
        owner = CATEGORY_OWNER.get(cat, "?")
        ws = wb.create_sheet(meta["title"][:30])  # sheet name max 31 chars
        ws.append([
            "题号", "题目", "答案", "验证状态", "解析", "证据路径",
            "证据 finding", "证据摘要", "来源角色", "置信度", "验证者",
            "验证时间", "验证备注", "更新时间",
        ])
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        for qid, qtext in meta["items"]:
            a = ans_idx.get((cat, qid), {})
            ans = a.get("answer", "") or "—"
            vs = (a.get("verification_status") or "unverified") if a.get("answer") else "未答"
            ev_id = a.get("evidence", "") or ""
            ev_summary = ""
            if ev_id and ev_id in findings_by_id:
                ev_summary = findings_by_id[ev_id].get("summary", "")
            ev_paths = a.get("evidence_path", []) or []
            row = [
                qid,
                qtext,
                ans,
                vs,
                a.get("analysis", "") or "",
                "\n".join(ev_paths),
                ev_id,
                ev_summary,
                a.get("source_role", ""),
                a.get("confidence", ""),
                a.get("verified_by", "") or "",
                a.get("verified_at", "") or "",
                a.get("verify_note", "") or "",
                a.get("updated", ""),
            ]
            ws.append(row)
            # 给验证状态格子上色
            cell = ws.cell(row=ws.max_row, column=4)
            if vs in status_fill:
                cell.fill = status_fill[vs]

        # 列宽
        widths = [8, 36, 28, 12, 60, 35, 12, 40, 16, 10, 14, 18, 30, 18]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[chr(64 + i)].width = w

        # 自动换行
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        # 解析列行高更高
        for r_idx in range(2, ws.max_row + 1):
            ws.row_dimensions[r_idx].height = 60

    # ── 跨角色 finding sheet ──
    cross = [f for f in findings if f.get("related_to")]
    if cross:
        ws_cross = wb.create_sheet("跨角色发现")
        ws_cross.append(["Finding ID", "来源", "类型", "摘要", "详情", "关联角色", "更新时间"])
        for cell in ws_cross[1]:
            cell.fill = header_fill
            cell.font = header_font
        for f in cross:
            related = ",".join(f.get("related_to", []))
            ws_cross.append([
                f.get("id", ""),
                f.get("from", ""),
                f.get("type", ""),
                f.get("summary", ""),
                f.get("detail", ""),
                related,
                f.get("updated", ""),
            ])
        for col_letter, width in [("A", 12), ("B", 18), ("C", 12), ("D", 50), ("E", 50), ("F", 30), ("G", 20)]:
            ws_cross.column_dimensions[col_letter].width = width

    wb.save(out_path)
    print(f"[+] Wrote {out_path}")


# ─────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--hub", default="http://127.0.0.1:8765",
                    help="Hub URL (default: http://127.0.0.1:8765)")
    ap.add_argument("--out", default=None,
                    help="Knowledge base directory (default: knowledge/competitions/2026FIC-团体赛/)")
    args = ap.parse_args()

    project_root = Path(__file__).resolve().parent.parent
    if args.out:
        kb_dir = Path(args.out)
    else:
        kb_dir = project_root / "knowledge" / "competitions" / "2026FIC-团体赛"
    kb_dir.mkdir(parents=True, exist_ok=True)

    print(f"[*] Hub:  {args.hub}")
    print(f"[*] KB:   {kb_dir}")

    state = fetch_hub_state(args.hub)
    print(f"[*] Hub state: findings={len(state['findings'])} "
          f"progress_roles={len(state['progress'])} "
          f"answer_categories={len(state['answers']) if state['answers'] else 0}")

    # MASTER_SHEET.md
    master_md = render_master_sheet(state)
    out_path = kb_dir / "MASTER_SHEET.md"
    out_path.write_text(master_md, encoding="utf-8")
    print(f"[+] Wrote {out_path} ({len(master_md)} chars)")

    # findings_snapshot.yaml
    snap_path = kb_dir / "findings_snapshot.yaml"
    snap = {
        "snapshot_time": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "hub": args.hub,
        "findings": state["findings"],
        "answers":  state["answers"],
        "progress": state["progress"],
        "session":  state["session"],
    }
    snap_path.write_text(
        yaml.safe_dump(snap, allow_unicode=True, sort_keys=False),
        encoding="utf-8",
    )
    print(f"[+] Wrote {snap_path}")

    # MASTER_SHEET.xlsx (Excel report - per user global rule: every report must have xlsx)
    xlsx_path = kb_dir / "MASTER_SHEET.xlsx"
    render_xlsx(state, xlsx_path)

    print("[+] Done.")


if __name__ == "__main__":
    main()
