"""
解析 usdt_transaction_ledger_70_records.xlsx, 找保险柜信息
"""
from pathlib import Path

SRC = Path(r"E:\ffffff-JIANCAI\2026FIC团体赛\case\binary\vhd_extracted\usdt_transaction_ledger_70_records.xlsx")
print(f"文件: {SRC}, 大小: {SRC.stat().st_size}")

with open(SRC, "rb") as f:
    head = f.read(8)
print(f"magic: {head.hex()}")

import openpyxl
wb = openpyxl.load_workbook(str(SRC))
print(f"\nSheets: {wb.sheetnames}")
for sn in wb.sheetnames:
    ws = wb[sn]
    print(f"\n=== Sheet: {sn} ({ws.max_row} rows × {ws.max_column} cols) ===")
    for r in range(1, min(ws.max_row + 1, 80)):
        row = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        row_str = " | ".join(str(v)[:60] for v in row)
        print(f"R{r}: {row_str}")

# 关键词搜
print("\n=== 全文搜 保险/safe/编号/密码/password ===")
for sn in wb.sheetnames:
    ws = wb[sn]
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if v and any(kw in str(v) for kw in ["保险", "safe", "编号", "密码", "password", "代码", "code", "金库"]):
                print(f"R{r}C{c}: {v}")

print("\n完成")
