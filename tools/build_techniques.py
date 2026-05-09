"""
批量创建 9 张通用技巧卡:
- dji_flightrecord_parse
- inet_ntoa_conversion
- sqlcipher_decrypt
- lxc_attach
- maccms_internals
- multichoice_enumeration
- veracrypt_manual_mount
- mp4_stco_encryption
- excel_xor_decrypt
- platform_format_dictionary (奖励, 第 10 张)
"""
from pathlib import Path
import yaml

KB = Path(r"E:\ffffff-JIANCAI\2026FIC团体赛\case\shared\knowledge_base")
TECH = KB / "techniques"
TECH.mkdir(parents=True, exist_ok=True)


TECHNIQUES = {
    "dji_flightrecord_parse": {
        "name": "DJI FlightRecord 解析",
        "category": "二进制日志",
        "when_to_use": [
            "题目涉及无人机飞行轨迹/经纬度/起降点/时间",
            "检材中找到 .txt 文件但 cat 是乱码 (DJI 加密二进制)",
            "DJI/FlightRecord/ 或 ~/DJI/Fly/Records/ 目录下文件",
        ],
        "detect_signal": [
            "文件 magic 是特定字节 (非纯文本)",
            "DJI 应用数据库内有 fly_record 关联",
            "题目关键词: DJI / 大疆 / 无人机 / 航拍 / 轨迹",
        ],
        "algorithm": (
            "DJI FlightRecord 是分层结构: header (元数据) + 多个 frame (按时间序列)\n"
            "每个 frame 有 osd (OSD 数据: lat/lon/alt/speed/...) 等子结构\n"
            "v13+ 版本 frame 数据被 keychain 加密, 需要 DJI 官方 API key 解密\n"
            "header/details 任何版本都不需要 key"
        ),
        "script": {
            "lang": "python",
            "code": (
                "from pydjirecord import DJILog\n"
                "import os\n\n"
                "data = open('FlightRecord_xxx.txt', 'rb').read()\n"
                "log = DJILog.from_bytes(data)\n"
                "print(f'version={log.version}, details={log.details}')\n\n"
                "api_key = os.environ.get('DJI_API_KEY')  # 申请: dji.com/cn/developer\n"
                "keychains = log.fetch_keychains(api_key) if log.version >= 13 else None\n"
                "for f in log.frames(keychains):\n"
                "    osd = f.osd\n"
                "    if osd and osd.latitude not in (None, 0.0):\n"
                "        print(osd.latitude, osd.longitude, osd.altitude)\n"
            ),
        },
        "seen_in": ["FIC2026_INITIAL_M_Q10"],
        "common_pitfalls": [
            "❌ 看到 .txt 后缀以为是文本",
            "❌ v13+ 没有 API key 时 frames() 抛异常",
            "❌ 经纬度 0.0/NaN 是无效采样, 要过滤",
            "❌ 反查行政区不准确 (用百度/高德 API, 别用 OpenStreetMap)",
        ],
        "related_techniques": ["gis_reverse_lookup"],
        "install": "pip install pydjirecord",
    },

    "inet_ntoa_conversion": {
        "name": "整数 IP ↔ IPv4 字符串转换",
        "category": "数据库 / 网络",
        "when_to_use": [
            "数据库表字段名带 _ip / ip_int / ipaddr 但值是整数",
            "看到 user_last_login_ip / user_reg_ip 是数字 (e.g. 859467171)",
            "Discuz / WordPress / maccms / phpBB 这类老 CMS 系统",
        ],
        "detect_signal": [
            "DESCRIBE 表显示该字段是 int unsigned (4字节)",
            "字段值是整数 (0 ~ 4294967295) 而不是 dotted notation",
        ],
        "algorithm": "IPv4 = a.b.c.d → 整数 = a*256^3 + b*256^2 + c*256 + d (大端)",
        "script": {
            "lang": "sql",
            "code": (
                "-- MySQL / MariaDB / TiDB\n"
                "SELECT inet_ntoa(user_last_login_ip) AS ip_str FROM mac_user;\n"
                "SELECT inet_aton('192.168.1.1') AS ip_int;  -- 反向\n\n"
                "-- PostgreSQL\n"
                "SELECT host(inet '192.168.1.1');\n\n"
                "-- Python\n"
                "import socket, struct\n"
                "ip_str = socket.inet_ntoa(struct.pack('!I', 859467171))\n"
                "ip_int = struct.unpack('!I', socket.inet_aton('51.43.21.163'))[0]\n"
            ),
        },
        "seen_in": ["FIC2026_INITIAL_S_Q15"],
        "common_pitfalls": [
            "❌ 直接读整数当 IP",
            "❌ 字节序搞反 (必须大端 '!I' / 网络字节序)",
        ],
        "related_techniques": ["maccms_internals"],
    },

    "sqlcipher_decrypt": {
        "name": "SQLCipher 加密 SQLite 解密",
        "category": "数据库 / 移动取证",
        "when_to_use": [
            "看到 .db / .sqlite 文件 sqlite3 报错 'file is not a database'",
            "Android IM 应用 (uuutalk / WeChat / WhatsApp / Telegram) 的聊天数据库",
            "iOS 应用类似情况",
        ],
        "detect_signal": [
            "file <db> 显示『data』而不是『SQLite 3.x database』",
            "前 16 字节不是 'SQLite format 3\\0'",
            "应用包名 + 数据库文件名通常包含 wk_ / e_ / encrypted_ 前缀",
        ],
        "algorithm": (
            "SQLCipher 默认用 PBKDF2-HMAC-SHA1/SHA256 派生 key, AES-256-CBC 加密\n"
            "密码通常硬编码在 APK / 二进制 / 用户输入派生 / 文件名本身 (本次比赛 uuutalk 用文件名 32 位 hex)"
        ),
        "script": {
            "lang": "bash",
            "code": (
                "# Linux (ubuntu): apt install sqlcipher\n"
                "sqlcipher wk_xxxxx.db <<EOF\n"
                "PRAGMA key = 'xxxxx';  -- 32位hex去掉wk_\n"
                ".tables\n"
                "SELECT * FROM message LIMIT 10;\n"
                ".dump\n"
                "EOF\n\n"
                "# Python\n"
                "from pysqlcipher3 import dbapi2 as sqlite\n"
                "conn = sqlite.connect('wk_xxxxx.db')\n"
                "c = conn.cursor()\n"
                "c.execute(\"PRAGMA key='xxxxx'\")\n"
                "c.execute('SELECT * FROM message LIMIT 5')\n"
            ),
        },
        "seen_in": ["FIC2026_INITIAL_M_Q4", "FIC2026_INITIAL_M_Q5", "FIC2026_INITIAL_M_Q6"],
        "common_pitfalls": [
            "❌ 用普通 sqlite3 cli (没编译 sqlcipher 支持)",
            "❌ 密码错误时 PRAGMA key 不会立刻报错, 要 .tables 才发现",
            "❌ KDF 迭代数版本差异 (sqlcipher 3 vs 4 默认不同, 用 PRAGMA cipher_compatibility = 3)",
        ],
        "related_techniques": [],
        "install": "apt install sqlcipher / pip install pysqlcipher3",
    },

    "lxc_attach": {
        "name": "LXC 容器内部访问",
        "category": "Linux 容器",
        "when_to_use": [
            "看到 ip a 有 lxcbr0 网桥",
            "/var/lib/lxc/<name>/config 存在",
            "容器内有目标服务 (如 TiDB/MySQL/特定应用)",
        ],
        "detect_signal": [
            "ps -aux 看到 lxc-start 或 lxc-monitor 进程",
            "/proc/1/cgroup 显示 0::/lxc/...",
            "lxc-ls -f 列出容器",
        ],
        "algorithm": "lxc 容器与宿主网络隔离, 必须 lxc-attach 进入或宿主路由到容器 IP",
        "script": {
            "lang": "bash",
            "code": (
                "# 列容器\n"
                "lxc-ls -f\n\n"
                "# 进入容器交互 shell\n"
                "lxc-attach -n mytidb\n\n"
                "# 一行执行命令\n"
                "lxc-attach -n mytidb -- bash -c 'ss -tlnp | grep 4000'\n\n"
                "# 复制文件 host ↔ container\n"
                "cp /tmp/x.txt /var/lib/lxc/mytidb/rootfs/tmp/\n\n"
                "# 看容器内 mysql/tidb\n"
                "lxc-attach -n mytidb -- mysql -h 127.0.0.1 -P 4000 -u root -e 'SHOW DATABASES'\n"
            ),
        },
        "seen_in": ["FIC2026_INITIAL_S_Q12", "FIC2026_INITIAL_S_Q13", "FIC2026_INITIAL_I_Q2"],
        "common_pitfalls": [
            "❌ 容器外直接 ping 容器 IP / mytidb hostname (网络隔离)",
            "❌ 把 docker 命令套到 lxc (不是 docker exec, 是 lxc-attach)",
            "❌ 容器没启动 lxc-attach 失败, 要 lxc-start -n <name> -d",
        ],
        "related_techniques": ["maccms_internals"],
    },

    "maccms_internals": {
        "name": "苹果 CMS (maccms) 数据结构",
        "category": "Web 应用 / 数据库",
        "when_to_use": ["题目涉及 maccms / 苹果 CMS / 视频网站后台"],
        "detect_signal": ["路径 /var/www/html/maccms / /www/wwwroot/maccms"],
        "algorithm": "maccms 标准结构, 表前缀 mac_, 主配置在 application/extra/maccms.php",
        "script": {
            "lang": "yaml",
            "code": (
                "files:\n"
                "  /application/extra/maccms.php:  # 主配置 (域名/ICP/SEO)\n"
                "  /application/database.php:     # 数据库连接\n"
                "  /template/<name>/info.ini:     # 模板配置 (被使用的模板)\n"
                "  /template/<name>/html/*.html:   # 具体模板文件\n\n"
                "tables:\n"
                "  mac_user:\n"
                "    user_id, user_name, user_pwd  # 密码 md5\n"
                "    user_reg_time (unix ts)\n"
                "    user_last_login_ip (int unsigned)  # ⚠ 需 inet_ntoa\n"
                "  mac_type:\n"
                "    type_id, type_name (中文), type_en (拼音 slug)\n"
                "  mac_vod:\n"
                "    vod_id, vod_name, vod_pic (封面图 URL/文件名)\n"
                "    vod_play_url, vod_time (上架时间)\n\n"
                "admin_login:\n"
                "  默认 /admin.php, 但常被改名 (本次为 user.php)\n"
                "  改名标志: maccms.php 中 ENTRANCE 常量\n"
            ),
        },
        "seen_in": [
            "FIC2026_INITIAL_S_Q5", "FIC2026_INITIAL_S_Q6", "FIC2026_INITIAL_S_Q7",
            "FIC2026_INITIAL_S_Q8", "FIC2026_INITIAL_S_Q9", "FIC2026_INITIAL_S_Q15",
            "FIC2026_INITIAL_I_Q2",
        ],
        "common_pitfalls": [
            "❌ 在 nginx 配置里找 ICP (实际在 maccms.php)",
            "❌ 答 type_name 的拼音 (实际有 type_en 字段)",
            "❌ 把『模板源』当 zip 包 (实际是 info.ini)",
        ],
        "related_techniques": ["inet_ntoa_conversion", "lxc_attach"],
    },

    "multichoice_enumeration": {
        "name": "多选题穷举验证",
        "category": "答题策略",
        "when_to_use": ["选择题 (尤其多选), 题目给 A/B/C/D/E 选项"],
        "detect_signal": ["题目末尾有 A.xxx B.xxx C.xxx 选项列表"],
        "algorithm": (
            "对每个选项独立验证, 用至少 2 个工具/数据源交叉确认\n"
            "选项中可能有: 真实存在的(选), 不存在的虚构项(不选), 装了但没运行的(看题意)"
        ),
        "script": {
            "lang": "bash",
            "code": (
                "# 文件系统多选\n"
                "lsblk -f      # 显示所有块设备的 FSTYPE\n"
                "blkid         # 显示已知文件系统标签\n"
                "df -T         # 已挂载分区 + 类型\n"
                "cat /proc/filesystems  # 内核支持列表\n\n"
                "# 数据库服务多选\n"
                "dpkg -l | grep -iE 'mysql|tidb|postgres|mariadb|sqlite|redis|mongo'\n"
                "systemctl list-units --type=service | grep -iE 'mysql|postgres|tidb'\n"
                "ss -tlnp | grep -E ':(3306|3307|4000|5432|6379|27017)'\n\n"
                "# 容器技术多选\n"
                "which docker lxc-ls podman containerd nerdctl\n"
                "ls /var/lib/{docker,lxc,containerd}/ 2>/dev/null\n"
            ),
        },
        "seen_in": [
            "FIC2026_INITIAL_S_Q16", "FIC2026_INITIAL_S_Q17",
            "FIC2026_INITIAL_M_Q11",
        ],
        "common_pitfalls": [
            "❌ 单选答 1 个漏其他 (题目明确多选时永远穷举)",
            "❌ 看到 GuessDB / FakeOS 这类虚构项不知排除",
            "❌ 只装 client 不算『安装服务』(必须 server 包)",
            "❌ 平台格式: A,B,D vs ABD vs A、B、D, 看历史 platform_confirmed",
        ],
        "related_techniques": ["platform_format_dictionary"],
    },

    "veracrypt_manual_mount": {
        "name": "VeraCrypt 容器手动挂载 (无火眼)",
        "category": "加密容器",
        "when_to_use": [
            "火眼自动解密失败但密码是对的",
            "Linux 环境, 没有火眼仿真",
            "需要批量/无 GUI 环境处理",
        ],
        "detect_signal": [
            "文件大小是整 MB/GB (10MB / 1GB), 内容看似随机",
            "火眼提示『需要 VeraCrypt 解密』",
            "文件无 magic 头 (前 64 字节是 VC 加密头, 看起来全是随机)",
        ],
        "algorithm": "VeraCrypt 用 AES/Twofish/Serpent + 多种 hash + 密码迭代派生 key",
        "script": {
            "lang": "bash",
            "code": (
                "# 命令行挂载 (Linux)\n"
                "veracrypt --text --mount <volume_file> /mnt/vc \\\n"
                "    --password='9ed2@99y8.com.cn' \\\n"
                "    --pim=0 --keyfiles='' --protect-hidden=no\n\n"
                "# 卸载\n"
                "veracrypt --text --dismount /mnt/vc\n\n"
                "# Windows GUI: 启动 VeraCrypt → Select File → Mount → 输入密码\n\n"
                "# 试探所有密码 (字典攻击)\n"
                "for pwd in $(cat passwords.txt); do\n"
                "    veracrypt --text --mount vc_file /mnt/vc \\\n"
                "        --password=\"$pwd\" --pim=0 --keyfiles='' --non-interactive 2>/dev/null \\\n"
                "        && echo \"FOUND: $pwd\" && break\n"
                "done\n"
            ),
        },
        "seen_in": ["FIC2026_INITIAL_C_Q8", "FIC2026_INITIAL_C_Q9", "FIC2026_INITIAL_C_Q10"],
        "common_pitfalls": [
            "❌ 密码错以为是火眼 bug, 实际可能是另一密码",
            "❌ pim 参数错误 (默认 0, 但有的设置非默认)",
            "❌ 没装 veracrypt cli (apt install veracrypt 或下载 .deb)",
        ],
        "related_techniques": ["cross_material_password"],
    },

    "mp4_stco_encryption": {
        "name": "MP4 stco 偏移加密",
        "category": "媒体文件加密",
        "when_to_use": [
            "mp4 文件存在但播放器无法播放 (报错 invalid frame)",
            "ffprobe 显示 metadata 正常但 chunks 错位",
        ],
        "detect_signal": [
            "ffprobe 报错 'Invalid offset' / 'EOF before chunk'",
            "二进制看 stco box 存在且 entry_count 合理",
            "知道有加密程序 (反汇编看到 stco 关键字)",
        ],
        "algorithm": (
            "MP4 box 结构: ftyp/moov/trak/mdia/minf/stbl/stco\n"
            "stco = sample table chunk offsets, 每条 4 字节 big-endian 是数据 chunk 在文件中的绝对偏移\n"
            "加密: 每个 offset += K (常见 K=1337/0x539, 0x10000)\n"
            "解密: 每个 offset -= K"
        ),
        "script": {
            "lang": "python",
            "code": (
                "import struct\n"
                "from pathlib import Path\n\n"
                "def decrypt_stco(path: Path, key: int = 1337, sub: bool = True):\n"
                "    \"\"\"sub=True 减 key 解密, False 加 key 加密\"\"\"\n"
                "    data = bytearray(path.read_bytes())\n"
                "    idx = data.find(b'stco')\n"
                "    if idx < 0:\n"
                "        return None\n"
                "    entry_count = struct.unpack('>I', data[idx+8:idx+12])[0]\n"
                "    for i in range(entry_count):\n"
                "        pos = idx + 12 + i*4\n"
                "        old = struct.unpack('>I', data[pos:pos+4])[0]\n"
                "        new = old - key if sub else old + key\n"
                "        if new < 0:\n"
                "            continue\n"
                "        data[pos:pos+4] = struct.pack('>I', new)\n"
                "    return bytes(data)\n"
            ),
        },
        "seen_in": ["FIC2026_INITIAL_C_Q9"],
        "common_pitfalls": [
            "❌ 用 little-endian unpack '<I' (mp4 是 big-endian)",
            "❌ 没考虑 64-bit offset (co64 box, 大文件用)",
            "❌ 改完 offset 没重写文件",
        ],
        "related_techniques": [],
    },

    "excel_xor_decrypt": {
        "name": "Excel 自定义 XOR-加 编码反算",
        "category": "应用层加密",
        "when_to_use": [
            "Excel 文件能打开但单元格内容是数字/乱码, 不是题目应有内容",
            "用户搜索历史有『如何加密 excel』",
            "tool 目录有加密脚本",
        ],
        "detect_signal": [
            "encoded 值范围奇怪 (e.g. 100000+ 而内容应是 6 位密码)",
            "找到加密脚本, 看到形如 (a^k)*N + (b^k)",
        ],
        "algorithm": (
            "通用结构: encoded = (((x + offset_x) ^ key_x) * multiplier) + ((y + offset_y) ^ key_y)\n"
            "反算: x = ((encoded // multiplier) ^ key_x) - offset_x\n"
            "      y = ((encoded % multiplier) ^ key_y) - offset_y"
        ),
        "script": {
            "lang": "python",
            "code": (
                "from openpyxl import load_workbook\n\n"
                "OFFSET = 100\n"
                "KEY = 85\n"
                "MULT = 1000\n\n"
                "def decode(encoded: int) -> tuple[int, int]:\n"
                "    x = ((encoded // MULT) ^ KEY) - OFFSET\n"
                "    y = ((encoded % MULT) ^ KEY) - OFFSET\n"
                "    return x, y\n\n"
                "wb = load_workbook('encrypted.xlsx')\n"
                "for ws in wb.worksheets:\n"
                "    for row in ws.iter_rows():\n"
                "        for cell in row:\n"
                "            if isinstance(cell.value, int):\n"
                "                print(cell.coordinate, decode(cell.value))\n"
            ),
        },
        "seen_in": ["FIC2026_INITIAL_C_Q10"],
        "common_pitfalls": [
            "❌ 试图用 john/excel-加密暴破 (不是标准加密)",
            "❌ 没找到加密脚本就放弃",
            "❌ 没注意到 OFFSET/KEY/MULT 是题目可变常量, 必须从加密脚本读",
        ],
        "related_techniques": ["browser_history_mining"],
    },

    "platform_format_dictionary": {
        "name": "比赛平台答案格式潜规则",
        "category": "答题策略",
        "when_to_use": ["所有题, 提交答案前必查"],
        "detect_signal": [
            "题目末尾【参考格式：xxx】",
            "题目说【不区分大小写/空格/全半角】但仍有未提到的格式",
        ],
        "algorithm": "早期收集 platform_confirmed 答案, 提取格式特征, 建字典",
        "script": {
            "lang": "yaml",
            "code": (
                "# 已知 FIC 平台规则字典\n"
                "telegram_group_id: '@FIC_2026'        # 必须带 @\n"
                "qq_group_id: '12345678'                # 不带 @\n"
                "wechat_group_id: 'wxid_xxxxx'\n\n"
                "date_formats:\n"
                "  - 'YYYY/M/D'      # 不补 0, 题目参考 '2000/1/1'\n"
                "  - 'YYYYMMDD'      # 紧凑, 题目参考 '20200101'\n"
                "  - 'YYYY-MM-DD'    # ISO, 题目参考 '2000-01-01'\n\n"
                "multichoice:\n"
                "  - 'A,B,D'         # 逗号分隔 (FIC 习惯)\n"
                "  - 'ABD'           # 紧凑 (DID-CTF 习惯)\n"
                "  - 'A、B、D'       # 中文顿号 (一些教育平台)\n\n"
                "url:\n"
                "  - 完整: 'https://example.com/path?q=1'\n"
                "  - 无 query: 'https://example.com/path'\n"
                "  - 仅 host: 'example.com'\n\n"
                "ip:\n"
                "  - IPv4: '192.168.1.1'\n"
                "  - 不缩写 0 (192.168.001.001 错)\n\n"
                "hash:\n"
                "  - 大写还是小写? 看题目参考格式样例字母大小写\n"
                "  - md5/sha1/sha256/sm3 长度: 32/40/64/64\n\n"
                "filename:\n"
                "  - 含扩展名? (5ab3e871....jpg vs 5ab3e871...)\n"
                "  - 路径? (/var/log/access.log vs access.log)\n"
            ),
        },
        "seen_in": [
            "FIC2026_INITIAL_I_Q1", "FIC2026_INITIAL_S_Q14",
            "FIC2026_INITIAL_M_Q11", "FIC2026_INITIAL_S_Q16", "FIC2026_INITIAL_S_Q17",
        ],
        "common_pitfalls": [
            "❌ 答 FIC_2026 缺 @",
            "❌ 日期补 0: 2026/04/15 vs 题目参考 2026/4/15",
            "❌ 多选用错分隔符",
            "❌ hash 大小写不匹配",
        ],
        "related_techniques": ["multichoice_enumeration"],
    },
}


for slug, t in TECHNIQUES.items():
    out = TECH / f"{slug}.yaml"
    if out.exists():
        # 不覆盖已有完整版本 (e.g. rsa_fermat)
        existing = yaml.safe_load(out.read_text(encoding="utf-8")) or {}
        if existing.get("seen_in"):  # 已有内容
            print(f"  ⊝ {slug}.yaml (已存在, 跳过)")
            continue
    
    # 标准字段
    data = {"slug": slug, **t}
    with out.open("w", encoding="utf-8") as f:
        yaml.safe_dump(data, f, allow_unicode=True, sort_keys=False, width=100)
    print(f"  ✓ {slug}.yaml")

print(f"\n已建 {len(TECHNIQUES)} 张技巧卡 → {TECH}")
