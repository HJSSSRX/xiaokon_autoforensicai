#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""把 questions_raw.txt 转成 questions.md + questions.xlsx + TASKS.md + phone inbox"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

BASE = Path("/mnt/e/项目/自动化取证/cases/2026FIC电子取证")

# 题目清单（人工解析，结构化）
QUESTIONS = [
    # (序号, 板块, 题目板块号, 题目文本, 分数, 检材, 分给, 优先级)
    # 计算机部分（Linux 桌面 - MAIN）
    ("C01", "计算机", 1, "分析计算机检材，操作系统版本号为", 9, "PC", "MAIN", "速通"),
    ("C02", "计算机", 2, "分析计算机检材，李安弘曾收到一份免费领取token的邮件的疑似钓鱼邮件，其发送用户邮箱为", 10, "PC", "MAIN", "中"),
    ("C03", "计算机", 3, "分析计算机检材，李安弘电脑中记录的黄金换现金的商家联系方式为", 11, "PC", "MAIN", "中"),
    ("C04", "计算机", 4, "分析计算机检材，推广设计图中的apk下载链接为", 13, "PC", "MAIN", "中"),
    ("C05", "计算机", 5, "分析计算机检材，李安弘电脑vpn软件开放的代理端口为", 10, "PC", "MAIN", "速通"),
    ("C06", "计算机", 6, "分析计算机检材，李安弘电脑中AI软件当前使用的模型类型为", 11, "PC", "MAIN", "中"),
    ("C07", "计算机", 7, "分析计算机检材，李安弘电脑中AI软件当前使用的模型apiKey为", 12, "PC", "MAIN", "中"),
    ("C08", "计算机", 8, "分析计算机检材，李安弘电脑中勒索软件提供的解密服务联系方式为", 11, "PC", "MAIN", "中"),
    ("C09", "计算机", 9, "分析计算机检材，李安弘电脑中记录的存放黄金的保险柜编号是", 13, "PC", "MAIN", "中"),
    ("C10", "计算机", 10, "分析计算机检材，李安弘电脑中记录的保险柜密码是", 13, "PC", "MAIN", "中"),
    # 手机部分（小米 MIUI - PHONE）
    ("M01", "手机", 1, "分析手机检材，该手机型号为", 9, "PHONE", "PHONE", "速通"),
    ("M02", "手机", 2, "分析手机检材，李安弘手机计划前往迪拜的日期是", 10, "PHONE", "PHONE", "中"),
    ("M03", "手机", 3, "分析手机检材，李安弘手机中与网站搭建人员沟通所使用的app安装日期为", 11, "PHONE", "PHONE", "中"),
    ("M04", "手机", 4, "分析手机检材，李安弘手机中与网站搭建人员沟通所使用的app，存放聊天数据的数据库为", 12, "PHONE", "PHONE", "中"),
    ("M05", "手机", 5, "分析手机检材，存放聊天数据的数据库的解密密码为", 13, "PHONE", "PHONE", "高"),
    ("M06", "手机", 6, "分析手机检材，李安弘购买云服务器商家的收款备用钱包地址为", 11, "PHONE", "PHONE", "中"),
    ("M07", "手机", 7, "分析手机检材，李安弘手机中给网站搭建人员第一次转账的交易hash前6位为", 10, "PHONE", "PHONE", "中"),
    ("M08", "手机", 8, "分析手机检材，手机中使用的AI软件李安弘主动向AI提问了几次", 11, "PHONE", "PHONE", "中"),
    ("M09", "手机", 9, "分析手机检材，李安弘手机使用的AI软件调用本地AI模型及版本为", 12, "PHONE", "PHONE", "中"),
    ("M10", "手机", 10, "分析手机检材，李安弘曾使用无人机航拍,分析其飞行轨迹，其在哪个县进行飞行", 10, "PHONE", "PHONE", "中"),
    ("M11", "手机", 11, "分析手机检材，李安弘最近安装了一个视频类APP，该APP声明了多个敏感权限用于收集用户隐私。请选择其中涉及用户隐私的敏感权限。", 11, "PHONE", "PHONE", "中"),
    ("M12", "手机", 12, "上述APP启动后会加载一个色情网站。请找出该APP当网络不可用时APP加载的本地离线页面路径。", 12, "PHONE", "PHONE", "中"),
    ("M13", "手机", 13, "上述APP将非法收集的用户隐私数据上传至远程服务器。上传地址在代码中经过编码处理。请找出编码方式，还原出完整的上传服务器URL。", 12, "PHONE", "PHONE", "高"),
    ("M14", "手机", 14, "该APP在本地创建了SQLite数据库存储收集到的用户信息。请分析代码，写出用于存储用户信息的表名", 13, "PHONE", "PHONE", "中"),
    ("M15", "手机", 15, "该APP的assets目录中存在一个加密配置文件config.dat。请解密该文件，写出其中的USDT钱包地址", 14, "PHONE", "PHONE", "高"),
    ("M16", "手机", 16, "该APP前端JS代码可以直接调用Android原生方法获取用户隐私数据。请分析暴露了哪些方法用于获取通讯录？", 12, "PHONE", "PHONE", "中"),
    ("M17", "手机", 17, "当主上传服务器不可达时，APP会获取备用服务器地址。请分析备用服务器的完整域名和端口", 14, "PHONE", "PHONE", "中"),
    # 服务器部分（Linux LVM - MAIN）
    ("S01", "服务器", 1, "该服务器主机操作系统版本为", 9, "SRV", "MAIN", "速通"),
    ("S02", "服务器", 2, "该服务器根分区硬盘的uuid号为", 11, "SRV", "MAIN", "速通"),
    ("S03", "服务器", 3, "该服务器中最新的docker镜像创建时间为", 10, "SRV", "MAIN", "中"),
    ("S04", "服务器", 4, "该服务器根分区快照路径为", 11, "SRV", "MAIN", "中"),
    ("S05", "服务器", 5, "该网站后台管理入口对应的文件名为", 11, "SRV", "MAIN", "中"),
    ("S06", "服务器", 6, "该网站设置的icp备案号为", 10, "SRV", "MAIN", "速通"),
    ("S07", "服务器", 7, "该网站设置的主域名为", 12, "SRV", "MAIN", "速通"),
    ("S08", "服务器", 8, "该网站分类3中，视频的拼音为", 12, "SRV", "MAIN", "中"),
    ("S09", "服务器", 9, "该站点设置页面中，被使用的前端模板来自于哪个源文件？", 13, "SRV", "MAIN", "中"),
    ("S10", "服务器", 10, "该网站的伪静态规则配置文件sm3值为", 12, "SRV", "MAIN", "中"),
    ("S11", "服务器", 11, "该网站关联的数据库的ip地址为", 11, "SRV", "MAIN", "中"),
    ("S12", "服务器", 12, "该网站数据库使用了哪一类容器技术", 10, "SRV", "MAIN", "速通"),
    ("S13", "服务器", 13, "运行在4000端口的备份数据库版本号为", 11, "SRV", "MAIN", "中"),
    ("S14", "服务器", 14, "新注册用户数量最多的日期为", 11, "SRV", "MAIN", "中"),
    ("S15", "服务器", 15, "马慧美最后一次登录该网站的ip为", 12, "SRV", "MAIN", "中"),
    ("S16", "服务器", 16, "以下哪个文件系统未被使用", 12, "SRV", "MAIN", "中"),
    ("S17", "服务器", 17, "该服务器安装了以下那些数据库服务", 13, "SRV", "MAIN", "速通"),
    # 互联网部分（MAIN，跨检材）
    ("I01", "互联网", 1, "售卖卡密的公开群组ID为", 13, "跨", "MAIN", "中"),
    ("I02", "互联网", 2, "备份数据库中视频图片的文件名为", 13, "SRV", "MAIN", "中"),
    ("I03", "互联网", 3, "ngrok提供的域名为", 14, "跨", "MAIN", "中"),
    # 二进制部分（U盘 + SampleVC.exe - MAIN）
    ("B01", "二进制", 1, "分析u盘检材，找到其中保存的加密程序SampleVC.exe，请给出这个exe程序的md5值？", 9, "USB", "MAIN", "速通"),
    ("B02", "二进制", 2, "分析SampleVC.exe，该程序编译的日期可能是什么？", 10, "USB", "MAIN", "速通"),
    ("B03", "二进制", 3, "分析SampleVC.exe，正确的密码是什么？", 14, "USB", "MAIN", "高"),
    ("B04", "二进制", 4, "分析u盘检材，利用SampleVC.exe解密U盘中被加密的文件，解密后的文件的后缀是什么？", 12, "USB", "MAIN", "高"),
    ("B05", "二进制", 5, "分析u盘检材，找到被加密的交易记录，统计李安弘虚拟币收款地址钱包总收款金额为", 14, "USB", "MAIN", "高"),
]

total_score = sum(q[4] for q in QUESTIONS)
main_score = sum(q[4] for q in QUESTIONS if q[6] == "MAIN")
phone_score = sum(q[4] for q in QUESTIONS if q[6] == "PHONE")

# === 1. questions.md ===
md = ["# FIC2026 题目清单\n",
      f"**总题数**：{len(QUESTIONS)} | **总分**：{total_score}\n",
      f"**MAIN**（计算机+服务器+互联网+二进制）：{sum(1 for q in QUESTIONS if q[6]=='MAIN')} 题 / {main_score} 分\n",
      f"**PHONE**（手机）：{sum(1 for q in QUESTIONS if q[6]=='PHONE')} 题 / {phone_score} 分\n\n",
      "---\n\n"]
current = None
for q in QUESTIONS:
    code, sect, num, text, score, evi, owner, prio = q
    if sect != current:
        md.append(f"\n## {sect}部分\n\n")
        current = sect
    md.append(f"### [{code}] Q{num} ({score}分) `{owner}` `{prio}`\n\n")
    md.append(f"> {text}\n\n")
(BASE / "questions.md").write_text("".join(md), encoding="utf-8")

# === 2. questions.xlsx ===
wb = Workbook()
ws = wb.active
ws.title = "FIC2026 题目"
headers = ["编号", "板块", "题号", "题目", "分值", "检材", "分给", "优先级", "状态", "答案", "证据路径", "备注"]
ws.append(headers)
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="4472C4")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

for q in QUESTIONS:
    code, sect, num, text, score, evi, owner, prio = q
    ws.append([code, sect, num, text, score, evi, owner, prio, "pending", "", "", ""])

# 列宽
widths = {"A": 8, "B": 10, "C": 6, "D": 70, "E": 6, "F": 6, "G": 8, "H": 8, "I": 10, "J": 30, "K": 30, "L": 20}
for col, w in widths.items():
    ws.column_dimensions[col].width = w
ws.row_dimensions[1].height = 30

# 板块染色
colors = {"计算机": "FFE699", "手机": "C6E0B4", "服务器": "F4B084", "互联网": "B4C7E7", "二进制": "F8CBAD"}
for row in range(2, len(QUESTIONS) + 2):
    sect = ws.cell(row=row, column=2).value
    if sect in colors:
        ws.cell(row=row, column=2).fill = PatternFill("solid", fgColor=colors[sect])
    # 自动换行
    ws.cell(row=row, column=4).alignment = Alignment(wrap_text=True, vertical="center")

# 第二个 sheet：分工汇总
ws2 = wb.create_sheet("分工汇总")
ws2.append(["板块", "题数", "总分", "分给", "占比%"])
for col in range(1, 6):
    cell = ws2.cell(row=1, column=col)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="4472C4")

stats = {}
for q in QUESTIONS:
    sect, score, owner = q[1], q[4], q[6]
    stats.setdefault(sect, {"count": 0, "score": 0, "owner": owner})
    stats[sect]["count"] += 1
    stats[sect]["score"] += score
for sect, d in stats.items():
    ws2.append([sect, d["count"], d["score"], d["owner"], f"{d['score']*100/total_score:.1f}%"])
ws2.append([])
ws2.append(["合计", len(QUESTIONS), total_score, "-", "100%"])
ws2.append([])
ws2.append(["MAIN", sum(1 for q in QUESTIONS if q[6]=='MAIN'), main_score, "MAIN", f"{main_score*100/total_score:.1f}%"])
ws2.append(["PHONE", sum(1 for q in QUESTIONS if q[6]=='PHONE'), phone_score, "PHONE", f"{phone_score*100/total_score:.1f}%"])
for col in "ABCDE":
    ws2.column_dimensions[col].width = 14

xlsx_path = BASE / "questions.xlsx"
wb.save(xlsx_path)

# === 3. 更新 TASKS.md ===
tasks_md = ["# FIC2026 题目分工表\n\n",
            f"**总题数**: {len(QUESTIONS)} | **总分**: {total_score}\n",
            f"**MAIN**（计算机+服务器+互联网+二进制）: {sum(1 for q in QUESTIONS if q[6]=='MAIN')} 题 / {main_score} 分\n",
            f"**PHONE**（手机）: {sum(1 for q in QUESTIONS if q[6]=='PHONE')} 题 / {phone_score} 分\n\n",
            "## 检材→板块固定分工\n\n",
            "| 检材 | 板块 | 状态 |\n|-----|------|------|\n",
            "| 检材1-计算机（Linux 桌面） | MAIN | /tmp/fic2026/pc/ewf/ewf1 |\n",
            "| 检材2-手机（小米 MIUI） | PHONE | tar 待解压 → /tmp/phone_extract |\n",
            "| 检材3-服务器（LVM 双盘） | MAIN | /dev/volum/root, /dev/root/data |\n",
            "| 检材4-U盘（NTFS+VeraCrypt） | MAIN | /tmp/fic2026/usb/ewf/ewf1 |\n\n",
            "## 题目分工（pending=待做, WIP=进行中, DONE=完成, SKIP=跳过, WAIT=等前置）\n\n",
            "| 编号 | 板块 | 分数 | 分给 | 优先 | 状态 | 题目 |\n",
            "|------|------|------|------|------|------|------|\n"]
for q in QUESTIONS:
    code, sect, num, text, score, evi, owner, prio = q
    short = text[:40] + ("..." if len(text) > 40 else "")
    tasks_md.append(f"| {code} | {sect} | {score} | {owner} | {prio} | pending | {short} |\n")

(BASE / "shared" / "TASKS.md").write_text("".join(tasks_md), encoding="utf-8")

# === 4. 给 PHONE 写 inbox ===
phone_qs = [q for q in QUESTIONS if q[6] == "PHONE"]
phone_md = [
    "# PHONE 窗口 inbox\n\n",
    "## 🚨 题目已下发（2026-04-25 13:05）\n\n",
    f"你的题目: **{len(phone_qs)} 题 / {phone_score} 分**（占总分 {phone_score*100/total_score:.0f}%）\n\n",
    "### 检材\n",
    "- 路径: `/mnt/e/ffffff-JIANCAI/2026FIC团体赛/检材2-手机.tar`\n",
    "- 4 GB / 38266 个条目 / **小米 MIUI** Android /data 分区\n",
    "- **第一步**: `mkdir -p /tmp/phone_extract && tar -xf /mnt/e/ffffff-JIANCAI/2026FIC团体赛/检材2-手机.tar -C /tmp/phone_extract` (慢，可后台跑)\n\n",
    "### 你的 17 题（全部）\n\n",
    "| 编号 | 题号 | 分 | 优先 | 题目 |\n|---|---|---|---|---|\n"
]
for q in phone_qs:
    code, sect, num, text, score, evi, owner, prio = q
    short = text.replace("\n", " ")
    phone_md.append(f"| {code} | M{num} | {score} | {prio} | {short} |\n")

phone_md.extend([
    "\n## 解题策略建议\n\n",
    "**速通题**（先做）:\n",
    "- M01 手机型号 → `cat /tmp/phone_extract/data/misc/recovery/ro.build.fingerprint`\n\n",
    "**有依赖链的题**（按顺序）:\n",
    "- M03(app安装日期) → M04(数据库文件名) → M05(解密密码)\n",
    "- M11(app权限) → M12(离线页面) → M13(URL编码) → M14(表名) → M15(config.dat) → M16(JS暴露方法) → M17(备用域名)\n",
    "  - **M11-M17 是同一个视频类 APP，需要先找到这个 APK 然后 jadx 反编译**\n\n",
    "**孤立题**:\n",
    "- M02 计划前往迪拜的日期 → 日历/聊天/邮件\n",
    "- M06,M07 钱包地址 → 微信/支付宝/聊天\n",
    "- M08,M09 AI 软件提问次数+本地模型 → 找 AI APP（如豆包/智谱/Ollama）\n",
    "- M10 无人机航拍县名 → 看 DJI APP 飞行日志或相册 EXIF\n\n",
    "## 共享发现\n",
    "- 在 `cases/2026FIC电子取证/shared/facts.md` 看 MAIN 的发现（可能有 IP/域名/账号互通）\n",
    "- 你发现的 IOC 也写到 facts.md（[PHONE] 标签）\n\n",
    "## 答案位置\n",
    "- `cases/2026FIC电子取证/shared/answers/M<NN>_phone_<简称>.md`\n",
    "- 严格 5 字段格式（题号+原文+答案+置信度+解析+不作弊声明）\n",
    "- 每题: `python scripts/notify_when_done.py 'M<NN> PHONE done'`\n",
    "- 每 5 题写 `cases/2026FIC电子取证/wp_batches/PHONE_MNN-MMM.md`\n"
])
(BASE / "shared" / "inbox" / "phone.md").write_text("".join(phone_md), encoding="utf-8")

print(f"✅ questions.md       {(BASE / 'questions.md').stat().st_size} bytes")
print(f"✅ questions.xlsx     {xlsx_path.stat().st_size} bytes")
print(f"✅ TASKS.md           updated")
print(f"✅ inbox/phone.md     updated")
print(f"")
print(f"题目: {len(QUESTIONS)} 道, 总分 {total_score}")
print(f"MAIN: {sum(1 for q in QUESTIONS if q[6]=='MAIN')} 题 / {main_score} 分")
print(f"PHONE: {sum(1 for q in QUESTIONS if q[6]=='PHONE')} 题 / {phone_score} 分")
