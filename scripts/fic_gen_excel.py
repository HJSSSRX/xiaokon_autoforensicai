#!/usr/bin/env python3
"""生成 FIC 2026 答题汇总 Excel"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ==================== 数据 ====================
# (题号, 板块, 分值, 题目简称, 题目原文, 答案, 置信度, 证据路径/命令, 备注)

answers = [
    # === 二进制 (B) ===
    ("B01", "MAIN-二进制", 9, "SampleVC.exe MD5",
     "计算U盘(检材2)中SampleVC.exe的MD5",
     "764789dd9c095d74b6b258cf0f7568b2", "高",
     "A:\\SampleVC.exe (检材2)",
     "md5sum SampleVC.exe"),

    ("B02", "MAIN-二进制", 10, "SampleVC.exe 编译日期",
     "获取U盘(检材2)中SampleVC.exe的编译日期（北京时间）",
     "2026-04-17 13:53:20 +08:00", "高",
     "A:\\SampleVC.exe PE TimeDateStamp",
     "readpe / PEiD，UTC 05:53:20 + 8h"),

    ("B03", "MAIN-二进制", 14, "SampleVC 密码", "", "未解", "—", "", "逆向分析 + 爆破待做"),
    ("B04", "MAIN-二进制", 12, "VC 加密文件解密", "", "未解", "—", "", "依赖 B03 密码"),
    ("B05", "MAIN-二进制", 14, "VC 内容", "", "未解", "—", "", "依赖 B04"),

    # === 计算机 (C) ===
    ("C01", "MAIN-计算机", 9, "PC 操作系统",
     "分析计算机检材，该计算机的操作系统版本为",
     "Deepin 23.1 (Deepin OS V23)", "高",
     "/tmp/pc_root/etc/os-release",
     "cat os-release → VERSION=\"23.1\""),

    ("C02", "MAIN-计算机", 10, "钓鱼邮件发件人",
     "分析计算机检材，发送给李安弘的钓鱼邮件的发件人邮箱为",
     "hf13338261292@outlook.com", "高",
     "/tmp/pc_data/root/.config/deepin/deepin-mail/ SQLite 邮件DB",
     "deepin-mail 数据库 messages 表"),

    ("C03", "MAIN-计算机", 11, "黄金换现金商家", "",
     "未解 (可能藏在 .et 加密文件或 VeraCrypt 容器)", "—", "", "跳过"),

    ("C04", "MAIN-计算机", 13, "推广图 APK 链接", "",
     "未解 (OCR 识别失败，需人眼看 /tmp/fic_extract/promo/*.jpeg)", "—",
     "/tmp/pc_data/root/图片/1-9.jpeg", "建议人工看图"),

    ("C05", "MAIN-计算机", 10, "VPN 应用端口",
     "分析计算机检材，该计算机上VPN应用的端口号",
     "主答: 7897 (mixed-port) | 候选: 9527 / 9097 / 7899", "中",
     "~/.local/share/io.github.clash-verge-rev.clash-verge-rev/config.yaml, verge.yaml",
     "verge_mixed_port=7897 最确证；9527 只在 Rust 符号哈希里巧合出现；9097 是 external-controller"),

    ("C06", "MAIN-计算机", 11, "AI 软件 模型类型",
     "分析计算机检材，李安弘电脑中AI软件当前使用的模型类型为",
     "DeepSeek (或更具体: DeepSeek-R1)", "高",
     "~/.local/share/deepin/uos-ai-assistant/db/basic (sqlite)",
     "llm 表 name='DeepSeek-试用账号'；二进制里有 r1-250120 / v3-250324"),

    ("C07", "MAIN-计算机", 12, "AI 模型 apiKey",
     "分析计算机检材，李安弘电脑中AI软件当前使用的模型apiKey为",
     "RFTYsCjYh/0F+p2gvEgpKtCQYHue1YQtOEU1GFS/I1Pu0HYpIWAvz/C2dNvxXi0FXPvMYE23hTewirnq1u63yg==", "中",
     "同 C06 sqlite 的 llm.account_proxy JSON 字段",
     "AES-128-ECB 加密的 base64；如需 sk- 明文需逆向"),

    ("C08", "MAIN-计算机", 11, "勒索软件解密联系", "", "未解", "—", "", "主机无 ransom/readme 痕迹"),
    ("C09", "MAIN-计算机", 13, "保险箱开箱", "", "未解 (.et 密码保护)", "—",
     "/root/文档/zhongyao/保险箱的秘密.et", "依赖 pass1/3/5/6/8 密码碎片"),
    ("C10", "MAIN-计算机", 13, "保险箱内容", "", "未解 (.et 密码保护)", "—", "", "依赖 C09"),

    # === 服务器 (S) ===
    ("S01", "MAIN-服务器", 9, "服务器 OS",
     "服务器操作系统版本为",
     "Debian GNU/Linux 13 (trixie)", "高",
     "/tmp/srv_root/@rootfs/etc/os-release",
     "VERSION=\"13 (trixie)\""),

    ("S02", "MAIN-服务器", 11, "根分区 UUID",
     "服务器根分区的UUID为",
     "3231e52f-5e15-44c4-b224-e29cb4201c0e", "高",
     "/etc/fstab 或 btrfs filesystem show",
     "subvol=@rootfs, btrfs RAID0"),

    ("S03", "MAIN-服务器", 10, "最新 docker 镜像创建时间",
     "服务器中最新的docker镜像创建时间为",
     "2026-04-16 07:15:50 UTC (北京 15:15:50)", "高",
     "/data/image/overlay2/repositories.json + image/metadata",
     "镜像 u22:latest"),

    ("S04", "MAIN-服务器", 11, "根分区快照路径",
     "服务器根分区的快照路径为",
     "/root/history", "高",
     "btrfs subvolume list → ID 257 /root/history",
     ""),

    ("S05", "MAIN-服务器", 11, "后台入口文件名",
     "服务器网站后台的入口文件名为",
     "user.php", "高",
     "/var/www/html/maccms10/ + nginx 配置",
     "maccms 被改名 admin→user"),

    ("S06", "MAIN-服务器", 10, "ICP 备案号",
     "服务器网站的ICP备案号为",
     "icp1919810", "高",
     "maccms 配置 + 模板 footer",
     ""),

    ("S07", "MAIN-服务器", 12, "主域名",
     "服务器网站的主域名为",
     "www.2026fic.forensix", "高",
     "nginx server_name + maccms site_url",
     ""),

    ("S08", "MAIN-服务器", 12, "分类3 拼音", "", "未解 (TiDB 容器未启动)", "—",
     "mac2.mac_type 表 type_id=3", "待启 LXC 容器查 db"),

    ("S09", "MAIN-服务器", 13, "站点设置模板",
     "服务器上的网站后台中\"站点设置\"页面前端模板来自于哪个源文件",
     "application/admin/view_new/system/config.html", "高",
     "System::config() + view_new/system/config.html",
     ""),

    ("S10", "MAIN-服务器", 12, "伪静态 sm3",
     "服务器伪静态配置的sm3哈希为",
     "5a6c86366041509a0a9a31b72a341372182e00f4c9df6a6ae34cc5e779639c54", "高",
     "nginx 伪静态规则 + sm3sum",
     ""),

    ("S11", "MAIN-服务器", 11, "数据库 IP",
     "服务器连接数据库的IP为",
     "10.0.3.100", "高",
     "maccms database.php + lxc-info mytidb",
     "LXC 容器 IP"),

    ("S12", "MAIN-服务器", 10, "容器技术",
     "服务器数据库使用的容器技术为",
     "LXC", "高",
     "/var/lib/lxc/mytidb/config",
     ""),

    ("S13", "MAIN-服务器", 13, "TiDB 版本", "", "未解 (容器未启动)", "—", "", "待启 LXC"),
    ("S14", "MAIN-服务器", 12, "注册用户最多日期", "", "未解 (需访问 mac2 数据库)", "—", "", ""),
    ("S15", "MAIN-服务器", 12, "马慧美登录 IP", "", "未解 (需访问 mac2 数据库)", "—", "", ""),

    ("S16", "MAIN-服务器", 12, "文件系统未被使用",
     "服务器中以下哪个文件系统未被使用",
     "候选: xfs (若选项含 ext4/xfs/btrfs/zfs)", "中",
     "/etc/fstab + /sbin/mkfs.*",
     "服务器已用 btrfs/vfat/swap/zfs/udf；无 mkfs.xfs；需看选项"),

    ("S17", "MAIN-服务器", 13, "数据库服务（多选）",
     "服务器安装了哪些数据库服务",
     "PostgreSQL + TiDB", "高",
     "dpkg 列表 + systemd services + LXC mytidb",
     "postgresql-17 本机装；TiDB 在 LXC；mariadb-client 只是 client"),

    # === 互联网 (I) ===
    ("I01", "MAIN-互联网", 13, "售卖卡密群组ID", "", "未解 (需访问邮件或数据库)", "—", "", ""),
    ("I02", "MAIN-互联网", 14, "备份视频图片名", "", "未解 (需访问 mac2 数据库)", "—", "", ""),

    ("I03", "MAIN-互联网", 14, "ngrok 域名",
     "ngrok提供的域名为",
     "blemish-junior-unengaged.ngrok-free.dev", "高",
     "/var/log/nginx/access.log.1 Referer 字段",
     "ngrok http 80 命令 + log 里 https://...ngrok-free.dev/"),

    # === 手机 (M) - PHONE 窗口负责 ===
    ("M01", "PHONE-手机", 9, "手机型号", "", "(PHONE 窗口已答，汇总待其报告)", "?", "", "让 PHONE 同步"),
    ("M02", "PHONE-手机", 10, "迪拜机票日期", "", "PHONE", "?", "", ""),
    ("M03", "PHONE-手机", 11, "沟通 app", "", "PHONE", "?", "", ""),
    ("M04", "PHONE-手机", 12, "聊天 db", "", "PHONE", "?", "", ""),
    ("M05", "PHONE-手机", 13, "聊天 db 密码", "", "PHONE", "?", "", ""),
    ("M06", "PHONE-手机", 12, "钱包地址", "", "PHONE", "?", "", ""),
    ("M07", "PHONE-手机", 14, "转账 hash", "", "PHONE", "?", "", ""),
    ("M08", "PHONE-手机", 11, "AI 问答数", "", "PHONE", "?", "", ""),
    ("M09", "PHONE-手机", 12, "本地 AI 模型", "", "PHONE", "?", "", ""),
    ("M10", "PHONE-手机", 12, "无人机县名", "", "PHONE", "?", "", ""),
    ("M11", "PHONE-手机", 11, "视频 app 权限", "", "PHONE", "?", "", ""),
    ("M12", "PHONE-手机", 10, "离线页面", "", "PHONE", "?", "", ""),
    ("M13", "PHONE-手机", 13, "上传 URL", "", "PHONE", "?", "", ""),
    ("M14", "PHONE-手机", 12, "userinfo 表", "", "PHONE", "?", "", ""),
    ("M15", "PHONE-手机", 14, "config.dat 钱包", "", "PHONE", "?", "", ""),
    ("M16", "PHONE-手机", 14, "jsbridge 通讯录", "", "PHONE", "?", "", ""),
    ("M17", "PHONE-手机", 13, "备份 endpoint", "", "PHONE", "?", "", ""),
]

# ==================== 生成 Workbook ====================
wb = Workbook()

# Sheet 1: 汇总
ws = wb.active
ws.title = "答题汇总"

header = ["题号", "板块", "分值", "题目简称", "题目原文", "答案", "置信度", "证据路径/命令", "备注"]
ws.append(header)

header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
header_font = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_align

body_font = Font(name="Microsoft YaHei", size=10)
body_align = Alignment(vertical="top", wrap_text=True)

conf_colors = {
    "高": "C6EFCE",
    "中": "FFEB9C",
    "低": "FFC7CE",
    "—": "D9D9D9",
    "?":  "E7E6E6",
}

for row in answers:
    ws.append(row)

thin = Side(border_style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for r in range(2, len(answers) + 2):
    for c in range(1, 10):
        cell = ws.cell(row=r, column=c)
        cell.font = body_font
        cell.alignment = body_align
        cell.border = border
    conf_cell = ws.cell(row=r, column=7)
    if conf_cell.value in conf_colors:
        conf_cell.fill = PatternFill(start_color=conf_colors[conf_cell.value], end_color=conf_colors[conf_cell.value], fill_type="solid")
        conf_cell.alignment = Alignment(horizontal="center", vertical="center")

widths = [8, 15, 6, 18, 40, 45, 8, 45, 30]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.freeze_panes = "A2"

# 行高
for r in range(2, len(answers) + 2):
    ws.row_dimensions[r].height = 50

# Sheet 2: 统计
ws2 = wb.create_sheet("得分统计")
ws2.append(["统计项", "数值"])

total = sum(row[2] for row in answers)
answered = [r for r in answers if r[5] and not r[5].startswith("未解") and not r[5].startswith("PHONE") and r[5] != "(PHONE 窗口已答，汇总待其报告)"]
answered_score = sum(r[2] for r in answered)
high_conf = [r for r in answered if r[6] == "高"]
high_conf_score = sum(r[2] for r in high_conf)
mid_conf = [r for r in answered if r[6] == "中"]
mid_conf_score = sum(r[2] for r in mid_conf)

rows_stat = [
    ("总题数", len(answers)),
    ("总分", total),
    ("已答题数", len(answered)),
    ("已答得分(乐观)", answered_score),
    ("高置信度题数", len(high_conf)),
    ("高置信度得分", high_conf_score),
    ("中置信度题数", len(mid_conf)),
    ("中置信度得分", mid_conf_score),
    ("MAIN 已答", len([r for r in answered if r[1].startswith("MAIN")])),
    ("MAIN 得分", sum(r[2] for r in answered if r[1].startswith("MAIN"))),
    ("PHONE 已答", "待 PHONE 汇报"),
]
for s in rows_stat:
    ws2.append(s)

for cell in ws2[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_align

for r in range(2, len(rows_stat) + 2):
    for c in range(1, 3):
        ws2.cell(row=r, column=c).font = body_font
        ws2.cell(row=r, column=c).border = border

ws2.column_dimensions["A"].width = 25
ws2.column_dimensions["B"].width = 20

# 保存
out = "/mnt/e/项目/自动化取证/cases/2026FIC电子取证/shared/answers/FIC2026_答题汇总.xlsx"
wb.save(out)
print(f"Saved: {out}")
print(f"总题数: {len(answers)}, 已答 {len(answered)} 题, 得分 {answered_score}/{total}")
